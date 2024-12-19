from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment
from django.db.models import Count

from .models import Robot


def get_last_week_robots_data():
    start_date = datetime.today() - timedelta(days=7)
    end_date = datetime.today()
    
    data = Robot.objects.filter(created__range=(start_date, end_date)) \
                        .values("model", "version").annotate(produced=Count("model")) \
                        .order_by("model")
    
    return data


def create_weekly_report():
    book = openpyxl.Workbook()
    
    data = get_last_week_robots_data()

    current_model = None

    for robot in data:
        model = robot.get("model")
        center_alignment = Alignment(horizontal='center', vertical='center')

        if model != current_model:
            if current_model is None:
                sheet = book[book.sheetnames[0]]
                sheet.title = model
                current_model = model
            else:
                current_model = model
                sheet = book.create_sheet(title=model)
            
            sheet["A1"] = "Модель"
            sheet["B1"] = "Версия"
            sheet["C1"] = "Количество за неделю"
            
            sheet.merge_cells("C1:E1")

            bold_font = Font(bold=True)
            sheet["A1"].font = bold_font
            sheet["B1"].font = bold_font
            sheet["C1"].font = bold_font

            sheet["C1"].alignment = center_alignment

            row = 2

        sheet[f"A{row}"] = robot.get("model")
        sheet[f"B{row}"] = robot.get("version")
        sheet[f"C{row}"] = robot.get("produced")
        sheet.merge_cells(f"C{row}:E{row}")

        sheet[f"C{row}"].alignment = center_alignment

        row += 1

    book.save("robots/media/weekly_report.xlsx")